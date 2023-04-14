import datetime
import re

import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication, QTreeView, QFileSystemModel, QVBoxLayout, QPushButton, QInputDialog, \
    QLineEdit, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime, date, timedelta
import numpy
import datetime
from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *


def get_search(keyword):
    import requests

    cookies = {
        'NNB': 'IZBJERH4DQMGI',
        'nx_ssl': '2',
        'ASID': '798f73fa0000018710b1765100000069',
        'CBI_SES': 'z/bB+dJz1Jkkcex3afxFCTWZmycH17VF26Hk8p3dZrxmLnIQlvB3kiZVhZtC4AQhePFU8xycJv+n2FHLfETNvFakwg0OzMWQdVSaewz4dU8+8qginjzS8gsFSwMJsgemdFb/rmbk7wODZJt14M3ArmmCyRb0iF+DcYO9eHZsaVKMOHpQIq3xiy/FXGH4cVqQpHQByNMZKNNKO6REpss8Alhda9LL47c3xtY4GwJo4UYy/9orLCajwLU/ZNsTIv2RrFVqyzU6B4DzYnmdyt/cZzXd53soc/wrDKMPrkV8M1+UiPZy3nE8LJ1kaLvUaOzpru8JTNzirIn6feKP1zNlcPtFTS38K8UVfKQB/qaa/NOFV2ysWke7RkSgPJNE/vW+C2CMub/675nSCqoGb7QrhTttA3dcMT5MAnv90Y+oY3keEVYjtqlZLCDSXcameHlvtZ9BhZjN73uaGO1zv/+24B2regAH8amL06gL7y3tYjY=',
        'nid_inf': '-1301193424',
        'NID_AUT': 'gcMRtUGaqkIrIA3N4IZVXgmQEpFaa+vBBt5JprursD/XiZFZRy2HsTY3RDlloFgj',
        'NID_JKL': '0mfi8F6SaYqMxcjF5XTtKxVb1QHBb4gW3twm3gFiIMM=',
        'NID_SES': 'AAABqcmc3umvOJ97RQfqYFIX5OhooCX01FhvhP/e0oxg1XoNNHM3MtR7Y0HDosmHyFMBwo6HaknFRuxnrcwKEWMANmtl10SQfSlVAKyL8UClA/ruslMHpLhYrChtY4Xr27wYgVUCFXqJ1aUWvC8N6nyd97hjhHoAZ4SspJzuXZARpPks9KfnIJh/wg5aSyaCCsyIdJ782xye8SqW/EJlZSK3nQ/1CMMt368mRztyZ8id+ghpqg+LbmYiDLwxoNh9WUKH5vhMnrne3b6d5XY2zaJZNC0/P+0gz0PJk4sl9vY3gWqmVOvUszWzNS5FRDa6A3H8/8dkcqGYM3E/bh47eZhWIAULRZuD+0NsGkHCSNZeI+v06f+FQdZGP8LLCTGzNZ86WMRRpPk71MdpBiBaU+tctOHZbbfcET1flOsg0W7iRkfm5lzD9u295heCEg+R6xM2HRO6k8HieOuE2umMvo7GjTssF5PqVB2NYicqSPL//h6Ug7Kqj2H0e0JQJt8AOUwUOhFeQJUzQ48sGt4XSMrCh6H3U3oGU4GWZjwaDHmpLPIR7J0f0Kj2Mcg3j+77o+Y5Rw==',
        '_naver_usersession_': '+uahvacQQVhvFP50vIf6+b0G',
        'page_uid': 'ivAG2wprvOsssnFal18ssssstqZ-501823',
        'CBI_CHK': '"r5V0mf9uRUZHZ/vmLGy3ez7f4/k4aqWXL5o03eN68foRF6vr8SfxFvlI+85vYVcKrwZpps2gKHNQX+ataQrIUGTL2zldbvohC/hpCypF3S0+KnoaTuvQS9tMcQY/s7igb7ne0ZUhx5rPSonkCoqSaznb2OvVu/eCQQBi01j5clQ="',
    }

    headers = {
        'authority': 'search.naver.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'cache-control': 'max-age=0',
        # 'cookie': 'NNB=IZBJERH4DQMGI; nx_ssl=2; ASID=798f73fa0000018710b1765100000069; CBI_SES=z/bB+dJz1Jkkcex3afxFCTWZmycH17VF26Hk8p3dZrxmLnIQlvB3kiZVhZtC4AQhePFU8xycJv+n2FHLfETNvFakwg0OzMWQdVSaewz4dU8+8qginjzS8gsFSwMJsgemdFb/rmbk7wODZJt14M3ArmmCyRb0iF+DcYO9eHZsaVKMOHpQIq3xiy/FXGH4cVqQpHQByNMZKNNKO6REpss8Alhda9LL47c3xtY4GwJo4UYy/9orLCajwLU/ZNsTIv2RrFVqyzU6B4DzYnmdyt/cZzXd53soc/wrDKMPrkV8M1+UiPZy3nE8LJ1kaLvUaOzpru8JTNzirIn6feKP1zNlcPtFTS38K8UVfKQB/qaa/NOFV2ysWke7RkSgPJNE/vW+C2CMub/675nSCqoGb7QrhTttA3dcMT5MAnv90Y+oY3keEVYjtqlZLCDSXcameHlvtZ9BhZjN73uaGO1zv/+24B2regAH8amL06gL7y3tYjY=; nid_inf=-1301193424; NID_AUT=gcMRtUGaqkIrIA3N4IZVXgmQEpFaa+vBBt5JprursD/XiZFZRy2HsTY3RDlloFgj; NID_JKL=0mfi8F6SaYqMxcjF5XTtKxVb1QHBb4gW3twm3gFiIMM=; NID_SES=AAABqcmc3umvOJ97RQfqYFIX5OhooCX01FhvhP/e0oxg1XoNNHM3MtR7Y0HDosmHyFMBwo6HaknFRuxnrcwKEWMANmtl10SQfSlVAKyL8UClA/ruslMHpLhYrChtY4Xr27wYgVUCFXqJ1aUWvC8N6nyd97hjhHoAZ4SspJzuXZARpPks9KfnIJh/wg5aSyaCCsyIdJ782xye8SqW/EJlZSK3nQ/1CMMt368mRztyZ8id+ghpqg+LbmYiDLwxoNh9WUKH5vhMnrne3b6d5XY2zaJZNC0/P+0gz0PJk4sl9vY3gWqmVOvUszWzNS5FRDa6A3H8/8dkcqGYM3E/bh47eZhWIAULRZuD+0NsGkHCSNZeI+v06f+FQdZGP8LLCTGzNZ86WMRRpPk71MdpBiBaU+tctOHZbbfcET1flOsg0W7iRkfm5lzD9u295heCEg+R6xM2HRO6k8HieOuE2umMvo7GjTssF5PqVB2NYicqSPL//h6Ug7Kqj2H0e0JQJt8AOUwUOhFeQJUzQ48sGt4XSMrCh6H3U3oGU4GWZjwaDHmpLPIR7J0f0Kj2Mcg3j+77o+Y5Rw==; _naver_usersession_=+uahvacQQVhvFP50vIf6+b0G; page_uid=ivAG2wprvOsssnFal18ssssstqZ-501823; CBI_CHK="r5V0mf9uRUZHZ/vmLGy3ez7f4/k4aqWXL5o03eN68foRF6vr8SfxFvlI+85vYVcKrwZpps2gKHNQX+ataQrIUGTL2zldbvohC/hpCypF3S0+KnoaTuvQS9tMcQY/s7igb7ne0ZUhx5rPSonkCoqSaznb2OvVu/eCQQBi01j5clQ="',
        'referer': 'https://www.naver.com/',
        'sec-ch-ua': '"Google Chrome";v="111", "Not(A:Brand";v="8", "Chromium";v="111"',
        'sec-ch-ua-arch': '"x86"',
        'sec-ch-ua-bitness': '"64"',
        'sec-ch-ua-full-version-list': '"Google Chrome";v="111.0.5563.149", "Not(A:Brand";v="8.0.0.0", "Chromium";v="111.0.5563.149"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-model': '""',
        'sec-ch-ua-platform': '"Windows"',
        'sec-ch-ua-platform-version': '"10.0.0"',
        'sec-ch-ua-wow64': '?0',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-site',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36',
    }

    params = {
        'where': 'nexearch',
        'sm': 'top_hty',
        'fbm': '1',
        'ie': 'utf8',
        'query': str(keyword),
    }

    response = requests.get('https://search.naver.com/search.naver', params=params, cookies=cookies, headers=headers)
    soup = BeautifulSoup(response.text, 'lxml')
    return soup
def get_urls(keyword):

    headers={"User-agent":'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36'}
    soup=get_search(keyword)

    type_a_urls = []
    type_b_urls = []
    type_c_urls = []
    urls = []
    names=[]
    urls_pre=[]


    all_smart_blocks=soup.find_all('section',attrs={'class':re.compile('sc_new sp_intent_block')})
    if len(all_smart_blocks)>=1:
        print("스마트블록많음",len(all_smart_blocks))
        for index,section in enumerate(all_smart_blocks):
            posting=""
            name=""
            try:
                title=section.find('div',attrs={'class':'title_area'}).get_text()
                print("text1:",title)
            except:
                title="없음"

            if title.find("인기 주제")>=0: #인기주제 URL 가져오기
                # text=section.find('strong',attrs={'class':'title'}).get_text()
                # print('text:', text)
                categorys=section.find_all('div',attrs={'class':'flick_bx'})
                for category in categorys: # 인기주제 내에서 칵 카드별 URL가져오기
                    url=category.find('a',attrs={'class':re.compile('popular_block_wrap')})
                    try:
                        name=category.find('div',attrs={'class':'dsc'}).get_text()
                    except:
                        name=""
                    if url==None:
                        break
                    url=url['href']
                    # print('url:', url)
                    # type_a_urls_pre.append(url)
                    urls_pre.append([url,name])

                #각 서칭 결과 확인하기
                for url_pre in urls_pre:
                    res=requests.get(url_pre[0],headers=headers)
                    soup=BeautifulSoup(res.text,'lxml')
                    ul_tag=soup.find('ul',attrs={'class':re.compile('keyword_challenge_list')})
                    # print('ul_tag:',ul_tag)
                    li_tags=ul_tag.find_all('li')
                    # print(len(li_tags))
                    for index,li_tag in enumerate(li_tags): # 각 카드 안에 들어가서 url 모두 가져오기
                        if index==5:
                            break
                        url=li_tag.find('a',attrs={'class':'title _intent_cross_collection_trigger'})['href']
                        name=li_tag.find('a',attrs={'class':'name'}).get_text()
                        posting = li_tag.find('a', attrs={'class': 'title _intent_cross_collection_trigger'}).get_text()
                        type_a_urls.append([index+1,name,url,posting,url_pre[1]])
                        # print('url:',url)
            else:
                type_c_url_element = []
                try:
                    text = section.find('strong', attrs={'class': 'intent_title'}).get_text()
                except:
                    print("VIEW타입아님")
                    continue
                print('text:', text)
                if text.find("인플루언서")>=0:
                    contents_cards=section.find_all('a',attrs={'class':'title _intent_cross_collection_trigger'})
                    for index,contents_card in enumerate(contents_cards):
                        url=contents_card['href']
                        name = section.find_all('a',attrs={'class':'name'})[index].get_text()
                        posting = section.find_all('a', attrs={'class': 'title _intent_cross_collection_trigger'})[index].get_text()
                        type_b_urls.append([index+1,name,url,posting])
                else:
                    contents_cards = section.find_all('a', attrs={'class': 'title _intent_cross_collection_trigger'})
                    for index,contents_card in enumerate(contents_cards):
                        url = contents_card['href']
                        name = section.find_all('a', attrs={'class': 'name'})[index].get_text()
                        posting = section.find_all('a', attrs={'class': 'title _intent_cross_collection_trigger'})[index].get_text()
                        type_c_urls.append([index+1,name,url,posting])
    else:
        print("스마트블록1개짜리")
        # print(soup.prettify())
        section=soup.find('section',attrs={'class':re.compile('sc_new sp_nreview')})
        contents_cards = section.find_all('li', attrs={'class': 'bx _svp_item'})
        for index, contents_card in enumerate(contents_cards):
            url = contents_card.find('a',attrs={'class':'api_txt_lines total_tit _cross_trigger'})['href']
            name = contents_card.find('a', attrs={'class': 'sub_txt sub_name'}).get_text()
            posting = contents_card.find('a', attrs={'class': 'api_txt_lines total_tit _cross_trigger'}).get_text()
            type_c_urls.append([index + 1, name, url, posting])


    print('type_a_urls:',type_a_urls)
    print('type_b_urls:',type_b_urls)
    print('type_c_urls:',type_c_urls)

    return type_a_urls,type_b_urls,type_c_urls
def load_excel(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    data_list = []
    for i in range(2, no_row + 1):
        name = ws.cell(row=i, column=1).value
        if name == "" or name == None:
            print('데이타 더 이상 없음')
            break
        keyword=ws.cell(row=i, column=2).value
        data = [name, keyword]
        data_list.append(data)
    print(data_list)
    return data_list

class Thread(QThread):
    cnt = 0
    user_signal = pyqtSignal(str)  # 사용자 정의 시그널 2 생성
    user_signal2 = pyqtSignal()  # 사용자 정의 시그널 2 생성

    def __init__(self, parent,fname,time_period):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.
        self.fname=fname
        self.time_period=time_period

    def run(self):
        print('엑셀생성')
        wb = openpyxl.Workbook()
        ws = wb.active
        column_name = ['제품', '키워드', '카테고리', '블로그명', '글제목', "URL", '키워드별상위노출개수', '순위','세부카테고리']
        ws.append(column_name)

        time_now_string = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        keyword_infos = load_excel(self.fname)  # 키워드가 져오기
        time_period = self.time_period

        for keyword_info in keyword_infos:
            count = 0
            time_now_check=datetime.datetime.now().strftime("%H시%M분%S초")
            text="상품 : {}, 키워드 : {} 검색중...({})".format(keyword_info[0],keyword_info[1],time_now_check)
            self.user_signal.emit(text)
            print("keyword:", keyword_info[1])
            try:
                type_a_urls, type_b_urls, type_c_urls = get_urls(keyword_info[1])  # 키워드에 대한 URL 모두 발췌하기
            except:
                print("브라우저 에러")
                continue

            search_list = ['위크나인', '워크나인', '위트나인', '워터나인']
            headers = {
                "User-agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36'}
            type_a_result = []
            type_b_result = []
            type_c_result = []

            for type_a_url in type_a_urls:
                res = requests.get(type_a_url[2], headers=headers)
                soup = BeautifulSoup(res.text, 'lxml')
                try:
                    info_iframe = soup.find('iframe', attrs={'id': 'mainFrame'})['src']
                except:
                    print("확인불가블로그")
                    continue
                real_url = 'https://blog.naver.com' + soup.find('iframe', attrs={'id': 'mainFrame'})['src']
                print('real_url:', real_url)
                res = requests.get(real_url, headers=headers)
                soup = BeautifulSoup(res.text, 'lxml')
                # print(soup.prettify())
                find_flag = False
                try:
                    contents = soup.find('div', attrs={'class': 'se-main-container'}).get_text()
                except:
                    print("에러인듯")
                    continue
                category = "스마트블록"
                for search_elem in search_list:
                    if contents.find(search_elem) >= 0:
                        data = [keyword_info[0], keyword_info[1], category, type_a_url[1], type_a_url[3], type_a_url[2],
                                "", type_a_url[0],type_a_url[4]]
                        count = count + 1
                        type_a_result.append(data)
                time.sleep(time_period)

            for type_b_url in type_b_urls:
                res = requests.get(type_b_url[2], headers=headers)
                soup = BeautifulSoup(res.text, 'lxml')
                try:
                    info_iframe = soup.find('iframe', attrs={'id': 'mainFrame'})['src']
                except:
                    print("확인불가블로그")
                    continue
                real_url = 'https://blog.naver.com' + soup.find('iframe', attrs={'id': 'mainFrame'})['src']
                print('real_url:', real_url)
                res = requests.get(real_url, headers=headers)
                soup = BeautifulSoup(res.text, 'lxml')
                # print(soup.prettify())
                find_flag = False
                try:
                    contents = soup.find('div', attrs={'class': 'se-main-container'}).get_text()
                except:
                    print("에러인듯")
                    continue
                category = "인플루언서탭"
                for search_elem in search_list:
                    if contents.find(search_elem) >= 0:
                        data = [keyword_info[0], keyword_info[1], category, type_b_url[1], type_b_url[3], type_b_url[2],
                                "", type_b_url[0]]
                        count = count + 1
                        type_b_result.append(data)
                time.sleep(time_period)

            for type_c_url in type_c_urls:
                res = requests.get(type_c_url[2], headers=headers)
                soup = BeautifulSoup(res.text, 'lxml')
                # print(soup.prettify())
                try:
                    info_iframe = soup.find('iframe', attrs={'id': 'mainFrame'})['src']
                except:
                    print("확인불가블로그")
                    continue
                real_url = 'https://blog.naver.com' + soup.find('iframe', attrs={'id': 'mainFrame'})['src']
                print('real_url:', real_url)
                res = requests.get(real_url, headers=headers)
                soup = BeautifulSoup(res.text, 'lxml')
                # print(soup.prettify())
                find_flag = False
                try:
                    contents = soup.find('div', attrs={'class': 'se-main-container'}).get_text()
                except:
                    print("에러인듯")
                    continue
                category = "VIEW"
                for search_elem in search_list:
                    if contents.find(search_elem) >= 0:
                        data = [keyword_info[0], keyword_info[1], category, type_c_url[1], type_c_url[3], type_c_url[2],
                                "", type_c_url[0]]
                        count = count + 1
                        type_c_result.append(data)
                time.sleep(time_period)

            print('type_a_result:', type_a_result)
            print('type_b_result:', type_b_result)
            print('type_c_result:', type_c_result)

            for type_a_result_elem in type_a_result:
                type_a_result_elem[6] = count
                ws.append(type_a_result_elem)

            for type_b_result_elem in type_b_result:
                type_b_result_elem[6] = count
                ws.append(type_b_result_elem)

            for type_c_result_elem in type_c_result:
                type_c_result_elem[6] = count
                ws.append(type_c_result_elem)
            wb.save('result_{}.xlsx'.format(time_now_string))
        self.user_signal2.emit()
    def stop(self):
        pass

class Example(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path = "C:"
        self.index = None
        self.setupUi(self)
        self.setSlot()
        self.show()
        self.time_period=float(self.lineEdit.text())
        QApplication.processEvents()

    def start(self):
        print('11')
        self.x = Thread(self,self.fname,self.time_period)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.user_signal2.connect(self.slot2)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()

    def slot1(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit.append(str(data1))

    def slot2(self):  # 사용자 정의 시그널1에 connect된 function
        QMessageBox.information(self, "완료창", "작업이 완료 되었습니다.")

    def find(self):
        print("find")
        self.fname=QFileDialog.getOpenFileName(self,"Open file",'./')[0]
        print(self.fname)
        self.lineEdit_2.setText(self.fname)

    def setSlot(self):
        pass

    def setIndex(self, index):
        pass

    def quit(self):
        QCoreApplication.instance().quit()


app = QApplication([])
ex = Example()
sys.exit(app.exec_())




# import datetime
# import re
#
# import requests
# from bs4 import BeautifulSoup
# import openpyxl
# import time
#
#
# wb=openpyxl.Workbook()
# ws=wb.active
# column_name=['제품','키워드','카테고리','블로그명','글제목',"URL",'키워드별상위노출개수','순위']
# ws.append(column_name)
#
# time_now_string=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
# keyword_infos=load_excel('keyword.xlsx') # 키워드가 져오기
# time_period=0.5
# for keyword_info in keyword_infos:
#     count=0
#     print("keyword:",keyword_info[1])
#     type_a_urls,type_b_urls,type_c_urls=get_urls(keyword_info[1]) #키워드에 대한 URL 모두 발췌하기
#
#     search_list=['위크나인','워크나인','위트나인','워터나인']
#     headers={"User-agent":'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36'}
#     type_a_result=[]
#     type_b_result=[]
#     type_c_result=[]
#
#     for type_a_url in type_a_urls:
#         res=requests.get(type_a_url[2],headers=headers)
#         soup=BeautifulSoup(res.text,'lxml')
#         try:
#             info_iframe=soup.find('iframe',attrs={'id':'mainFrame'})['src']
#         except:
#             print("확인불가블로그")
#             continue
#         real_url='https://blog.naver.com'+soup.find('iframe',attrs={'id':'mainFrame'})['src']
#         print('real_url:',real_url)
#         res = requests.get(real_url, headers=headers)
#         soup = BeautifulSoup(res.text, 'lxml')
#         # print(soup.prettify())
#         find_flag=False
#         try:
#             contents=soup.find('div',attrs={'class':'se-main-container'}).get_text()
#         except:
#             print("에러인듯")
#             continue
#         category="스마트블록"
#         for search_elem in search_list:
#             if contents.find(search_elem)>=0:
#                 data=[keyword_info[0],keyword_info[1],category,type_a_url[1],type_a_url[3],type_a_url[2],"",type_a_url[0]]
#                 count=count+1
#                 type_a_result.append(data)
#         time.sleep(time_period)
#
#     for type_b_url in type_b_urls:
#         res=requests.get(type_b_url[2],headers=headers)
#         soup=BeautifulSoup(res.text,'lxml')
#         try:
#             info_iframe=soup.find('iframe',attrs={'id':'mainFrame'})['src']
#         except:
#             print("확인불가블로그")
#             continue
#         real_url='https://blog.naver.com'+soup.find('iframe',attrs={'id':'mainFrame'})['src']
#         print('real_url:',real_url)
#         res = requests.get(real_url, headers=headers)
#         soup = BeautifulSoup(res.text, 'lxml')
#         # print(soup.prettify())
#         find_flag=False
#         try:
#             contents=soup.find('div',attrs={'class':'se-main-container'}).get_text()
#         except:
#             print("에러인듯")
#             continue
#         category="인플루언서탭"
#         for search_elem in search_list:
#             if contents.find(search_elem)>=0:
#                 data=[keyword_info[0],keyword_info[1],category,type_b_url[1],type_b_url[3],type_b_url[2],"",type_b_url[0]]
#                 count = count + 1
#                 type_b_result.append(data)
#         time.sleep(time_period)
#
#     for type_c_url in type_c_urls:
#         res=requests.get(type_c_url[2],headers=headers)
#         soup=BeautifulSoup(res.text,'lxml')
#         # print(soup.prettify())
#         try:
#             info_iframe=soup.find('iframe',attrs={'id':'mainFrame'})['src']
#         except:
#             print("확인불가블로그")
#             continue
#         real_url='https://blog.naver.com'+soup.find('iframe',attrs={'id':'mainFrame'})['src']
#         print('real_url:',real_url)
#         res = requests.get(real_url, headers=headers)
#         soup = BeautifulSoup(res.text, 'lxml')
#         # print(soup.prettify())
#         find_flag=False
#         try:
#             contents=soup.find('div',attrs={'class':'se-main-container'}).get_text()
#         except:
#             print("에러인듯")
#             continue
#         category="VIEW"
#         for search_elem in search_list:
#             if contents.find(search_elem)>=0:
#                 data=[keyword_info[0],keyword_info[1],category,type_c_url[1],type_c_url[3],type_c_url[2],"",type_c_url[0]]
#                 count = count + 1
#                 type_c_result.append(data)
#         time.sleep(time_period)
#
#     print('type_a_result:',type_a_result)
#     print('type_b_result:',type_b_result)
#     print('type_c_result:',type_c_result)
#
#     for type_a_result_elem in type_a_result:
#         type_a_result_elem[6]=count
#         ws.append(type_a_result_elem)
#
#     for type_b_result_elem in type_b_result:
#         type_b_result_elem[6] = count
#         ws.append(type_b_result_elem)
#
#     for type_c_result_elem in type_c_result:
#         type_c_result_elem[6] = count
#         ws.append(type_c_result_elem)
#     wb.save('result_{}.xlsx'.format(time_now_string))




